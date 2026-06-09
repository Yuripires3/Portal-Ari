#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Gera data/indicadores-consolidado.json a partir de data/indicadores.xlsx.
Dados estáticos (2021 até mai/2026) — executar apenas se o Excel for atualizado.
"""

from __future__ import annotations

import json
import math
import re
import unicodedata
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
EXCEL_PATH = ROOT / "data" / "indicadores.xlsx"
JSON_PATH = ROOT / "data" / "indicadores-consolidado.json"

MESES = list(range(1, 13))

OPERADORAS_POR_ANO: dict[int, list[str]] = {
    2021: [
        "Unimed Rio",
        "ASSIM SAUDE",
        "Integral Saude",
        "HealthMed",
        "Hapvida NotreDame",
        "CONSOLIDADO",
    ],
    2022: [
        "Unimed Rio",
        "ASSIM SAUDE",
        "Integral Saude",
        "HealthMed",
        "Hapvida NotreDame",
        "Amil",
        "Infinity Doctors",
        "CONSOLIDADO",
    ],
    2023: [
        "Unimed Rio",
        "ASSIM SAUDE",
        "Integral Saude",
        "HealthMed",
        "Hapvida NotreDame",
        "Amil",
        "Klini Saude",
        "blue.",
        "CONSOLIDADO",
        "Infinity Doctors",
    ],
    2024: [
        "Unimed Rio",
        "ASSIM SAUDE",
        "Integral Saude",
        "HealthMed",
        "Hapvida NotreDame",
        "Amil",
        "Klini Saude",
        "blue.",
        "Leve Saude",
        "NOVA SAUDE",
        "CONSOLIDADO",
        "AESP Odonto",
    ],
    2025: [
        "Unimed Rio",
        "ASSIM SAUDE",
        "SEGUROS Unimed",
        "Leve Saude",
        "NOVA SAUDE",
        "blue.",
        "Hapvida NotreDame",
        "Oplan",
        "HealthMed",
        "SAUDE ONIX",
        "CONSOLIDADO",
        "Amil",
        "Integral Saude",
        "AESP Odonto",
        "Klini Saude",
    ],
    2026: [
        "Unimed Rio",
        "ASSIM SAUDE",
        "SEGUROS Unimed",
        "Leve Saude",
        "NOVA SAUDE",
        "blue.",
        "Hapvida NotreDame",
        "Oplan",
        "HealthMed",
        "SAUDE ONIX",
        "MedSenior",
        "CONSOLIDADO",
        "Amil",
        "Integral Saude",
        "AESP Odonto",
    ],
}

# Operadoras somadas no CONSOLIDADO 2026 (fórmulas do Excel — sem AESP Odonto).
CONSOLIDADO_SOMA_OPERADORAS_2026 = [
    "Unimed Rio",
    "ASSIM SAUDE",
    "SEGUROS Unimed",
    "Leve Saude",
    "NOVA SAUDE",
    "blue.",
    "Hapvida NotreDame",
    "Oplan",
    "HealthMed",
    "SAUDE ONIX",
    "MedSenior",
    "Amil",
    "Integral Saude",
]

# Correções manuais (print oficial) — não regravar o Excel (openpyxl apaga cache de fórmulas).
AJUSTES_INDICADORES: dict[int, dict[str, dict[str, dict[str, float | None]]]] = {
    2026: {
        "MedSenior": {
            "base_saude": {"1": None, "2": 5, "3": 15, "4": 18, "5": 22, "6": None},
            "base_vidas": {"1": None, "2": 5, "3": 15, "4": 18, "5": 22, "6": None},
            "vidas_canceladas": {"1": None, "2": None, "3": None, "4": None, "5": None, "6": None},
            "outros": {"1": None, "2": 2, "3": 2, "4": None, "5": None, "6": None},
            "faturamento_emitido": {"1": None, "2": 5502, "3": 18780, "4": 18623, "5": 38151, "6": None},
            "faturamento_recebido": {"1": None, "2": 5130, "3": 17499, "4": 18623, "5": 35307, "6": None},
            "inadimplencia": {"1": 0, "2": 0.0675, "3": 0.0682, "4": 0, "5": 0.0745, "6": None},
            "vendas": {"1": None, "2": 5, "3": 10, "4": 3, "5": 4, "6": None},
            "ticket_medio": {"1": None, "2": 1834, "3": 1565, "4": 1242, "5": 1734, "6": None},
            "pct_cancelamento": {"1": 0, "2": 0, "3": 0, "4": 0, "5": 0, "6": 0},
        },
    },
}

CHAVES_SOMA_CONSOLIDADO = [
    "meta_orcada",
    "base_dental",
    "base_saude",
    "vidas_canceladas",
    "retencao",
    "cancel_inadimplencia",
    "cancel_solicitacao_cliente",
    "cancel_solicitado_ops",
    "falecimento",
    "outros",
    "faturamento_emitido",
    "faturamento_recebido",
    "vendas",
    "comissao_concessionarias",
    "bonificacao_corretores_supervisores",
]

LABEL_MAP = {
    "meta orcada": "meta_orcada",
    "base": "base_vidas",
    "base vidas": "base_vidas",
    "base dental": "base_dental",
    "base saude": "base_saude",
    "vidas canceladas": "vidas_canceladas",
    "migracao assim>assim (a partir de agosto)": "migracao_assim_assim",
    "migracao assim>outras operadoras (a partir de agosto)": "migracao_assim_outras",
    "migracao caberj>assim (a partir de agosto)": "migracao_caberj_assim",
    "migracao caberj>outras operadoras (a partir de agosto)": "migracao_caberj_outras",
    "total migracao": "total_migracao",
    "cancelamento liquido": "cancelamento_liquido",
    "retencao": "retencao",
    "% cancelamento": "pct_cancelamento",
    "cancel. por inadimplencia": "cancel_inadimplencia",
    "cancel. solicitacao cliente": "cancel_solicitacao_cliente",
    "cancel. solicitado ops": "cancel_solicitado_ops",
    "exclusao de dependente": "exclusao_dependente",
    "exclusao de dependente (a partir de maio)": "exclusao_dependente",
    "falecimento": "falecimento",
    "falecimento (a partir de maio)": "falecimento",
    "obito": "falecimento",
    "outros": "outros",
    "outros (a partir de maio)": "outros",
    "faturamento orcado": "faturamento_orcado",
    "faturamento emitido": "faturamento_emitido",
    "faturamento recebido": "faturamento_recebido",
    "inadimplencia": "inadimplencia",
    "inadimplencia do fechamento do mes": "inadimplencia",
    "vendas": "vendas",
    "ticket medio": "ticket_medio",
    "comissao concessionarias": "comissao_concessionarias",
    "bonificacao corretores/supervisores": "bonificacao_corretores_supervisores",
}


def norm_label(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = unicodedata.normalize("NFD", text)
    text = "".join(c for c in text if unicodedata.category(c) != "Mn")
    text = re.sub(r"\s+", " ", text)
    return text.lower()


def parse_val(value):
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
        if text in ("-", "#REF!", "#N/A", ""):
            return None
        if text.startswith("="):
            return None
        text = re.sub(r"[R$\s%]", "", text)
        if "," in text and "." in text:
            text = text.replace(".", "").replace(",", ".")
        elif "," in text:
            text = text.replace(",", ".")
        try:
            return float(text)
        except ValueError:
            return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
            return None
        return float(value)
    return None


def merge_meses_celulas(
    anterior: dict[str, float | None],
    novo: dict[str, float | None],
) -> dict[str, float | None]:
    """Mescla linhas duplicadas (sub-tabela 2022): prioriza valor novo, exceto 0 sobre dado existente."""
    resultado: dict[str, float | None] = {}
    for m in MESES:
        ms = str(m)
        val_novo = novo.get(ms)
        val_ant = anterior.get(ms)
        if val_novo is None:
            resultado[ms] = val_ant
        elif val_ant is None:
            resultado[ms] = val_novo
        elif val_novo == 0 and val_ant not in (None, 0):
            resultado[ms] = val_ant
        else:
            resultado[ms] = val_novo
    return resultado


def parse_block(ws, header_row: int, next_header: int | None) -> dict:
    end = (next_header - 1) if next_header else header_row + 45
    rows: dict[str, dict[str, float | None]] = {}
    for row in range(header_row + 1, end + 1):
        raw = ws.cell(row, 5).value
        if raw is None:
            continue
        label = norm_label(raw)
        # Alguns blocos 2022 têm sub-tabela (faturamento) após um 2º cabeçalho "Indicador"
        if label == "indicador":
            continue
        key = LABEL_MAP.get(label)
        if not key:
            continue
        meses = {str(m): parse_val(ws.cell(row, 6 + i).value) for i, m in enumerate(MESES)}
        if key in rows:
            rows[key] = merge_meses_celulas(rows[key], meses)
        else:
            rows[key] = meses
    return rows


def merge_indicators(
    a: dict[str, dict[str, float | None]],
    b: dict[str, dict[str, float | None]],
) -> dict[str, dict[str, float | None]]:
    result = {key: dict(meses) for key, meses in a.items()}
    for key, meses_b in b.items():
        if key not in result:
            result[key] = dict(meses_b)
            continue
        merged = {}
        for m in MESES:
            ms = str(m)
            val_b = meses_b.get(ms)
            val_a = result[key].get(ms)
            merged[ms] = val_b if val_b is not None else val_a
        result[key] = merged
    return result


def ano_da_aba(nome: str) -> int:
    digits = "".join(c for c in nome if c.isdigit())
    return 2000 + int(digits[-2:])


def obter_valor_mes(indicadores: dict, key: str, mes: int) -> float | None:
    ms = str(mes)
    if key == "base_vidas":
        bruto = indicadores.get("base_vidas", {}).get(ms)
        if bruto is not None:
            return bruto
        saude = indicadores.get("base_saude", {}).get(ms)
        dental = indicadores.get("base_dental", {}).get(ms)
        if saude is None and dental is None:
            return None
        return (saude or 0) + (dental or 0)
    if key not in indicadores:
        return None
    return indicadores[key].get(ms)


def aplicar_ajustes(operadoras: list[dict], ano: int) -> None:
    ajustes_ano = AJUSTES_INDICADORES.get(ano)
    if not ajustes_ano:
        return
    for bloco in operadoras:
        nome = bloco["operadora"]
        if nome not in ajustes_ano:
            continue
        indicadores = bloco.setdefault("indicadores", {})
        for chave, meses in ajustes_ano[nome].items():
            destino = indicadores.setdefault(chave, {})
            for mes_str, valor in meses.items():
                if valor is None:
                    destino.pop(mes_str, None)
                else:
                    destino[mes_str] = valor


def recalcular_consolidado_2026(operadoras: list[dict]) -> None:
    nomes = set(CONSOLIDADO_SOMA_OPERADORAS_2026)
    por_nome = {o["operadora"]: o for o in operadoras}
    consolidado = next((o for o in operadoras if o.get("tipo") == "consolidado"), None)
    if not consolidado:
        return

    indicadores: dict[str, dict[str, float | None]] = {chave: {} for chave in CHAVES_SOMA_CONSOLIDADO}
    indicadores["base_vidas"] = {}
    indicadores["pct_cancelamento"] = {}
    indicadores["inadimplencia"] = {}
    indicadores["ticket_medio"] = {}

    for mes in MESES:
        ms = str(mes)
        totais: dict[str, float | None] = {}
        for chave in CHAVES_SOMA_CONSOLIDADO:
            soma: float | None = None
            for nome in CONSOLIDADO_SOMA_OPERADORAS_2026:
                op = por_nome.get(nome)
                if not op:
                    continue
                valor = obter_valor_mes(op["indicadores"], chave, mes)
                if valor is not None:
                    soma = (soma or 0) + valor
            totais[chave] = soma
            if soma is not None:
                indicadores[chave][ms] = soma

        base_vidas = totais.get("base_saude")
        if base_vidas is not None or totais.get("base_dental") is not None:
            base_vidas = (totais.get("base_saude") or 0) + (totais.get("base_dental") or 0)
            indicadores["base_vidas"][ms] = base_vidas

        vidas_canceladas = totais.get("vidas_canceladas")
        if base_vidas and base_vidas > 0 and vidas_canceladas is not None:
            indicadores["pct_cancelamento"][ms] = vidas_canceladas / base_vidas

        fat_emitido = totais.get("faturamento_emitido")
        fat_recebido = totais.get("faturamento_recebido")
        if fat_emitido and fat_emitido > 0 and fat_recebido is not None:
            indicadores["inadimplencia"][ms] = 1 - (fat_recebido / fat_emitido)

        if base_vidas and base_vidas > 0 and fat_emitido is not None:
            indicadores["ticket_medio"][ms] = fat_emitido / base_vidas

    consolidado["indicadores"] = indicadores


def processar_ano(ws, ano: int, headers: list[int]) -> list[dict]:
    nomes = OPERADORAS_POR_ANO.get(ano, [])
    operadoras: list[dict] = []

    for index, header_row in enumerate(headers):
        # 2021: blocos 6 e 7 do Excel são metades do CONSOLIDADO QV
        if ano == 2021 and index == 6:
            continue
        if ano == 2021 and index == 5:
            ind_a = parse_block(ws, headers[5], headers[6])
            ind_b = parse_block(ws, headers[6], None)
            indicadores = merge_indicators(ind_a, ind_b)
            operadoras.append(
                {"operadora": "CONSOLIDADO", "tipo": "consolidado", "indicadores": indicadores}
            )
            continue

        # 2026: bloco 16 do Excel está vazio (legado Samp/Klini)
        if ano == 2026 and index >= len(nomes):
            continue

        # 2022: blocos 9 e 11 do Excel — consolidado (bloco 10 é lixo #REF)
        if ano == 2022 and index in (8, 9):
            continue
        if ano == 2022 and index == 7:
            ind_a = parse_block(ws, headers[7], headers[8])
            ind_b = parse_block(ws, headers[9], None)
            indicadores = merge_indicators(ind_a, ind_b)
            operadoras.append(
                {"operadora": "CONSOLIDADO", "tipo": "consolidado", "indicadores": indicadores}
            )
            continue

        next_header = headers[index + 1] if index + 1 < len(headers) else None
        indicadores = parse_block(ws, header_row, next_header)
        if not indicadores:
            continue

        nome = nomes[index] if index < len(nomes) else f"Operadora {index + 1}"
        tipo = (
            "consolidado"
            if nome.upper() in ("CONSOLIDADO", "QV TOTAL")
            else "operadora"
        )
        operadoras.append({"operadora": nome, "tipo": tipo, "indicadores": indicadores})

    return operadoras


def main() -> None:
    if not EXCEL_PATH.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {EXCEL_PATH}")

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    resultado = {"fonte": str(EXCEL_PATH.name), "anos": {}}

    for sheet_name in wb.sheetnames:
        ano = ano_da_aba(sheet_name)
        ws = wb[sheet_name]
        headers: list[int] = []
        for row in range(1, ws.max_row + 1):
            if ws.cell(row, 5).value == "Indicador" and ws.cell(row, 6).value == "Jan":
                headers.append(row)

        operadoras = processar_ano(ws, ano, headers)
        aplicar_ajustes(operadoras, ano)
        if ano == 2026:
            recalcular_consolidado_2026(operadoras)
        resultado["anos"][str(ano)] = {"operadoras": operadoras}
        print(f"{ano}: {len(operadoras)} blocos — {[o['operadora'] for o in operadoras]}")

    JSON_PATH.parent.mkdir(parents=True, exist_ok=True)
    with JSON_PATH.open("w", encoding="utf-8") as f:
        json.dump(resultado, f, ensure_ascii=False)
    print(f"Gerado: {JSON_PATH} ({JSON_PATH.stat().st_size:,} bytes)")


if __name__ == "__main__":
    main()
